# 分词
import jieba
import re
from openpyxl import load_workbook
from openpyxl import Workbook


def stop_word():
    stop_word = []
    fd = open("./stopwords/baidu_stopwords.txt", "r", encoding='utf-8')
    for line in fd.readlines():
        stop_word += str(line).strip('\n').split(',')
    return stop_word
    # print(stop_word)


pattern = re.compile(r'[\u4e00-\u9fa5]')  # 过滤非中文字符


def word_cut(list1):
    jieba.load_userdict('user_dict.txt')
    seg_list = []
    for i in range(len(list1)):
        words_list = jieba.lcut(list1[i], cut_all=False)
        for word in words_list:
            if pattern.match(word) and word not in stop_word():
                seg_list.append(word)
    # print(seg_list)
    return seg_list


if __name__ == "__main__":
    wb1 = load_workbook("selected_data.xlsx")
    result = Workbook()
    ws1 = wb1["Sheet"]
    resultSheet = result["Sheet"]
    max_row = ws1.max_row
    text_list = []
    word_dict = {}
    for m in range(1, max_row + 1):
        weibo_cell_index = 'C' + str(m)
        temp = ws1[weibo_cell_index].value
        text_list.append(temp)
    # print(text_list)
    word_list = word_cut(text_list)
    for word in word_list:
        if word not in word_dict:
            word_dict[word] = 1
        else:
            word_dict[word] += 1
##### 将字典中的词和词频存入excel
    max_row = len(word_dict)
    key_list = list(word_dict.keys())
    value_list = list(word_dict.values())
    count = 0
    for m in range(1, max_row + 1):
        i = '%s%d' % ('A', m)
        j = '%s%d' % ('B', m)
        resultSheet[i]=key_list[count]
        resultSheet[j]=value_list[count]
        count+=1
    result.save("word_and_frequency.xlsx")
    print(word_dict.keys())
