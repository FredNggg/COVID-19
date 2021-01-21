#通过词表和停用词表筛选所需的微博
from openpyxl  import Workbook
from openpyxl  import load_workbook
def key_word():
    key_word=[]
    fd = open( "key_word.txt", "r",encoding='utf-8' )
    for line in fd.readlines():
        key_word+=str(line).strip('\n').split(',')
    return key_word

# def select_data(list):

if __name__ == "__main__":
    wb1 = load_workbook("data_01.xlsx")
    wb2 = Workbook()
    ws1 = wb1["Sheet"]
    ws2 = wb2["Sheet"]
    max_row = ws1.max_row
    max_column = ws1.max_column
    count = 0
    word_list = key_word()
    for m in range(1, max_row + 1):
        weibo_cell_index = 'C' + str(m)
        for word in word_list:
            if word in ws1[weibo_cell_index].value:
                count = count + 1
                for n in range(97, 97 + max_column):  # chr(97)='a'
                    n = chr(n)  # ASCII字符
                    i = '%s%d' % (n, m)
                    j = '%s%d' % (n, count)# 单元格编号
                    cell1 = ws1[i].value  # 获取data单元格数据
                    ws2[j].value = cell1  # 赋值到test单元格
            break
    wb2.save('selected_data.xlsx')


