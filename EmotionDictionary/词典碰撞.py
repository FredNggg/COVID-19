from openpyxl import load_workbook
from openpyxl import Workbook


def emotion_dict():  ##将excel中的情绪词典转化为字典
    emotion_dict = {}
    wb1 = load_workbook('心态词典.xlsx')
    ws1 = wb1["Sheet1"]
    max_row = ws1.max_row
    for m in range(1, max_row + 1):
        i = '%s%d' % ('A', m)
        j = '%s%d' % ('B', m)
        emotion_dict[ws1[i].value] = int(ws1[j].value)
    return emotion_dict


def tag_in_excel():  ##将情感标签附到词频excel后
    result = Workbook()
    emo_dict = emotion_dict()
    resultSheet = result["Sheet"]
    for i in range(1, 5):
        filename = 'stage' + str(i) + '_comments_frequency.xlsx'
        wb1 = load_workbook(filename)
        ws1 = wb1["Sheet"]
        max_row = ws1.max_row
        for m in range(1, max_row + 1):
            a = '%s%d' % ('A', m)
            c = '%s%d' % ('C', m)
            ws1[c].value = emo_dict[ws1[a].value]
        wb1.save('stage' + str(i) + '_comments_frequency_with_tag.xlsx')

tag_in_excel()